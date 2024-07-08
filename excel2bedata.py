from sys import argv
from docxtpl import DocxTemplate
from openpyxl import load_workbook

# 测试阶段，要求接收一个地址参数，并在这个地址下做点什么
script, first = argv
# script是本文件的名字，first是本文件的名字+参数，组成的列表
print(script)
print(argv[1])
# 被调用地址
path = argv[1]
# 生成一个文件夹，测试用
# os.mkdir(path)

# 测试地址
# path = 'C:\\Users\\Administrator\\Desktop'
dic = {
}

# 根据地址读取为字典，方便融合
wb = load_workbook(path + '\\字段.xlsm')
table = wb.active

dic['申请编号'] = str(int(table.cell(3, 3).value))
dic['管理单位'] = table.cell(4, 3).value
dic['业务类别'] = table.cell(5, 3).value
dic['客户编号'] = str(int(table.cell(6, 3).value))
dic['客户名称'] = table.cell(7, 3).value
dic['客户地址'] = table.cell(8, 3).value
dic['行业类别'] = table.cell(9, 3).value
dic['用电类别'] = table.cell(10, 3).value
dic['联系人'] = table.cell(11, 3).value
dic['联系电话'] = str(int(table.cell(12, 3).value))
dic['申请容量'] = str(int(table.cell(13, 3).value))
dic['核定容量'] = table.cell(14, 3).value
dic['变电站'] = table.cell(15, 3).value
dic['线路'] = table.cell(16, 3).value
dic['下线点'] = table.cell(17, 3).value
dic['计量方式'] = table.cell(18, 3).value
dic['PT变比'] = table.cell(19, 3).value
dic['CT变比'] = table.cell(20, 3).value
dic['电价类别'] = table.cell(21, 3).value
dic['是否分时'] = table.cell(22, 3).value
dic['是否力调'] = table.cell(23, 3).value
dic['基本电费'] = table.cell(24, 3).value
dic['力调标准'] = table.cell(25, 3).value

# 补充备用字段
dic['备用字段甲'] = table.cell(30, 3).value
dic['备用字段乙'] = table.cell(31, 3).value
dic['备用字段丙'] = table.cell(32, 3).value
dic['备用字段丁'] = table.cell(33, 3).value
dic['备用字段戊'] = table.cell(34, 3).value
dic['备用字段戌'] = table.cell(35, 3).value

# 添加供电方案所需要的特殊内容
if u'杆' in dic[u'下线点']:
    dic[u'产权分界'] = '下线的附杆柱上开关设备负荷侧与用户电缆连接点处'
else:
    dic[u'产权分界'] = '用户下线电缆连接点处'

if dic['计量方式'] == '高供低计':
    dic['电压互感器'] = ''
    dic['接线方式'] = '三相四线'
else:
    dic['电压互感器'] = '电压互感器变比为' + dic['PT变比'] + ',0.2级;'
    dic['接线方式'] = '三相三线'

# 仅仅用+描述变压器台数，太单一，需要加入乘号
if '+' in dic['核定容量'] or '*' in dic['核定容量']:
    trans = (dic['核定容量']).split('+')
    transcount = 0
    # 在循环中使用变量需要先声明，否则就属于在循环中反复声明了。
    tstr = ''  # 描述变压器的字符串声明
    for tran in trans:
        # 再用乘号分割
        tra = tran.split('*')
        tstr = tstr + '、' + tra[0] + '台' + tra[1] + '千伏安变压器'
        transcount = transcount + int(tra[0])
    # 顿号不好处理，直接字符切割掉
    dic['变压器'] = tstr[1:]
    # print(dic['变压器'])
    dic['变压器数量'] = transcount

print(dic)

# 开启融合，并保存到指定位置
obj = DocxTemplate('C:\\template\\验收资料模板.docx')
obj.render(dic)
pathsave = path + '\\' + dic['申请编号'] + dic['客户名称'] + dic['业务类别'] + str(dic['申请容量']) + '验收资料.docx'
obj.save(pathsave)
